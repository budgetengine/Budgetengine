"""
Módulo de Exportação Excel Profissional COMPLETO v5
Budget Engine - Relatórios para Clientes
VERSÃO CONSOLIDADA E PROFUNDA
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Any


class EstilosExcel:
    """Estilos profissionais para Excel"""
    
    AZUL_ESCURO = '1F4E79'
    AZUL_MEDIO = '2E75B6'
    AZUL_CLARO = '5B9BD5'
    VERDE = '70AD47'
    VERDE_ESCURO = '375623'
    AMARELO = 'FFC000'
    LARANJA = 'ED7D31'
    VERMELHO = 'C00000'
    CINZA_ESCURO = '404040'
    CINZA_MEDIO = '808080'
    CINZA_CLARO = 'D9D9D9'
    BRANCO = 'FFFFFF'
    FUNDO_AZUL = 'DDEBF7'
    FUNDO_VERDE = 'E2EFDA'
    FUNDO_AMARELO = 'FFF2CC'
    FUNDO_VERMELHO = 'FCE4D6'
    
    @classmethod
    def borda_fina(cls):
        return Border(
            left=Side(style='thin', color=cls.CINZA_CLARO),
            right=Side(style='thin', color=cls.CINZA_CLARO),
            top=Side(style='thin', color=cls.CINZA_CLARO),
            bottom=Side(style='thin', color=cls.CINZA_CLARO)
        )
    
    @classmethod
    def borda_total(cls):
        return Border(
            top=Side(style='medium', color=cls.AZUL_ESCURO),
            bottom=Side(style='double', color=cls.AZUL_ESCURO)
        )


class ExcelBudgetExporter:
    """Exportador de Budget Completo para Excel - Versão Consolidada"""
    
    MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", 
             "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    
    def __init__(self, motor, cliente_nome: str = None, filial_nome: str = None):
        self.motor = motor
        # Usar parâmetros passados ou buscar do motor
        self.cliente_nome = cliente_nome or getattr(motor, 'cliente_nome', None) or "Cliente"
        self.filial_nome = filial_nome or getattr(motor, 'filial_nome', None) or "Filial"
        self.tipo_relatorio = getattr(motor, 'tipo_relatorio', 'Filial')
        self.wb = Workbook()
        self.estilos = EstilosExcel
        self._calcular_dados()
    
    def _calcular_dados(self):
        """Pré-calcula todos os dados necessários"""
        self.motor.calcular_receita_bruta_total()
        self.motor.calcular_deducoes_total()
        self.dre = self.motor.calcular_dre()
        self.fc = self.motor.calcular_fluxo_caixa()
        self.pe_anual = self.motor.calcular_pe_anual()
        self.ocupacao = self.motor.calcular_ocupacao_anual()
        self.tdabc = self.motor.calcular_tdabc_anual()
        self.dividendos = self.motor.calcular_dividendos()
        self.simples = self.motor.calcular_simples_nacional_anual()
    
    def _aplicar_estilo_cabecalho(self, cell):
        cell.font = Font(name='Calibri', size=10, bold=True, color=self.estilos.BRANCO)
        cell.fill = PatternFill('solid', fgColor=self.estilos.AZUL_MEDIO)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.estilos.borda_fina()
    
    def _aplicar_estilo_total(self, cell):
        cell.font = Font(name='Calibri', size=10, bold=True, color=self.estilos.AZUL_ESCURO)
        cell.fill = PatternFill('solid', fgColor=self.estilos.CINZA_CLARO)
        cell.border = self.estilos.borda_total()
    
    def _aplicar_estilo_resultado(self, cell, positivo=True):
        cell.font = Font(name='Calibri', size=10, bold=True, color=self.estilos.BRANCO)
        cell.fill = PatternFill('solid', fgColor=self.estilos.VERDE if positivo else self.estilos.VERMELHO)
        cell.border = self.estilos.borda_fina()
    
    def _criar_titulo_secao(self, ws, row: int, texto: str, col_fim: int = 15):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_fim)
        cell = ws.cell(row=row, column=2)
        cell.value = texto
        cell.font = Font(name='Calibri', size=12, bold=True, color=self.estilos.BRANCO)
        cell.fill = PatternFill('solid', fgColor=self.estilos.AZUL_ESCURO)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 22
        return row + 1
    
    def _criar_cabecalho_mensal(self, ws, row: int, col_inicio: int = 2, incluir_av: bool = False):
        headers = [""] + self.MESES + ["TOTAL"]
        if incluir_av:
            headers.append("AV%")
        
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=col_inicio + i)
            cell.value = header
            self._aplicar_estilo_cabecalho(cell)
        
        ws.row_dimensions[row].height = 20
        return row + 1
    
    def _escrever_linha(self, ws, row: int, conta: str, valores: List[float], 
                        nivel: int = 0, is_total: bool = False, is_resultado: bool = False,
                        receita_ref: float = None, is_percent: bool = False):
        # Nome da conta
        cell = ws.cell(row=row, column=2)
        prefixo = "  " * nivel
        cell.value = f"{prefixo}{conta}"
        cell.border = self.estilos.borda_fina()
        
        if is_resultado:
            positivo = sum(valores) >= 0
            self._aplicar_estilo_resultado(cell, positivo)
        elif is_total:
            self._aplicar_estilo_total(cell)
        elif nivel == 0:
            cell.font = Font(name='Calibri', size=10, bold=True, color=self.estilos.AZUL_ESCURO)
            cell.fill = PatternFill('solid', fgColor=self.estilos.FUNDO_AZUL)
        else:
            cell.font = Font(name='Calibri', size=10, color=self.estilos.CINZA_ESCURO)
        
        # Valores mensais
        for i, valor in enumerate(valores):
            cell = ws.cell(row=row, column=3 + i)
            cell.value = valor
            cell.border = self.estilos.borda_fina()
            
            if is_percent:
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0;(#,##0);"-"'
            
            if is_resultado:
                self._aplicar_estilo_resultado(cell, valor >= 0)
            elif is_total:
                self._aplicar_estilo_total(cell)
            elif nivel == 0:
                cell.font = Font(name='Calibri', size=10, bold=True)
                cell.fill = PatternFill('solid', fgColor=self.estilos.FUNDO_AZUL)
            else:
                if valor < 0 and not is_percent:
                    cell.font = Font(name='Calibri', size=10, color=self.estilos.VERMELHO)
            
            cell.alignment = Alignment(horizontal='right')
        
        # Total
        total = sum(valores) if not is_percent else (sum(valores) / len(valores) if valores else 0)
        cell = ws.cell(row=row, column=15)
        cell.value = total
        cell.number_format = '0.0%' if is_percent else '#,##0;(#,##0);"-"'
        cell.alignment = Alignment(horizontal='right')
        
        if is_resultado:
            self._aplicar_estilo_resultado(cell, total >= 0)
        elif is_total:
            self._aplicar_estilo_total(cell)
        else:
            cell.border = self.estilos.borda_fina()
            cell.font = Font(name='Calibri', size=10, bold=True)
        
        # AV%
        if receita_ref and not is_percent:
            cell = ws.cell(row=row, column=16)
            cell.value = total / receita_ref if receita_ref != 0 else 0
            cell.number_format = '0.0%'
            cell.border = self.estilos.borda_fina()
            cell.alignment = Alignment(horizontal='right')
        
        return row + 1

    def _set_larguras_padrao(self, ws, col_descricao: int = 35):
        """Define larguras padrão das colunas"""
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = col_descricao
        for i in range(12):
            ws.column_dimensions[get_column_letter(3+i)].width = 11
        ws.column_dimensions['O'].width = 13
        ws.column_dimensions['P'].width = 8

    # =========================================================================
    # ABA 1: CAPA PROFISSIONAL
    # =========================================================================
    def criar_capa(self):
        ws = self.wb.active
        ws.title = "Capa"
        
        # Configurar larguras
        ws.column_dimensions['A'].width = 3
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['H'].width = 3
        
        # Configurar alturas para espaçamento
        for r in range(1, 45):
            ws.row_dimensions[r].height = 18
        
        # =====================================================================
        # CABEÇALHO - Faixa azul escura
        # =====================================================================
        for r in range(2, 6):
            for c in range(2, 8):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill('solid', fgColor=self.estilos.AZUL_ESCURO)
        
        # Título principal
        ws.merge_cells('B3:G3')
        cell = ws['B3']
        cell.value = "ORÇAMENTO EMPRESARIAL"
        cell.font = Font(name='Calibri', size=28, bold=True, color=self.estilos.BRANCO)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('B4:G4')
        cell = ws['B4']
        cell.value = "Exercício 2026"
        cell.font = Font(name='Calibri', size=16, color=self.estilos.BRANCO)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # =====================================================================
        # INFORMAÇÕES DO CLIENTE
        # =====================================================================
        ws.row_dimensions[8].height = 30
        ws.merge_cells('B8:G8')
        cell = ws['B8']
        cell.value = self.cliente_nome.upper()
        cell.font = Font(name='Calibri', size=24, bold=True, color=self.estilos.AZUL_ESCURO)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('B9:G9')
        cell = ws['B9']
        cell.value = self.filial_nome
        cell.font = Font(name='Calibri', size=14, color=self.estilos.CINZA_MEDIO)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Linha divisória
        for c in range(2, 8):
            ws.cell(row=11, column=c).fill = PatternFill('solid', fgColor=self.estilos.AZUL_CLARO)
        ws.row_dimensions[11].height = 4
        
        # =====================================================================
        # KPIs PRINCIPAIS - Cards visuais
        # =====================================================================
        row = 13
        
        # Calcular dados
        receita = sum(self.dre.get('Receita Bruta Total', [0]*12))
        ebitda = sum(self.dre.get('EBITDA', [0]*12))
        lucro = sum(self.dre.get('Resultado Líquido', [0]*12))
        sessoes = sum(self.pe_anual.meses[m].total_sessoes for m in range(12))
        dividendos = sum(self.dividendos.get('cronograma', [0]*12))
        
        receita_liq = sum(self.dre.get('Receita Líquida', [0]*12))
        margem_ebitda = ebitda / receita_liq if receita_liq > 0 else 0
        margem_liq = lucro / receita_liq if receita_liq > 0 else 0
        
        pe_medio = sum(self.pe_anual.meses[m].pe_contabil for m in range(12)) / 12
        ms_media = sum(self.pe_anual.meses[m].margem_seguranca_pct for m in range(12)) / 12
        ocup_media = sum(self.ocupacao.meses[m].taxa_ocupacao_sala for m in range(12)) / 12
        
        # Título da seção
        ws.merge_cells(f'B{row}:G{row}')
        cell = ws.cell(row=row, column=2)
        cell.value = "INDICADORES-CHAVE DE DESEMPENHO"
        cell.font = Font(name='Calibri', size=12, bold=True, color=self.estilos.AZUL_ESCURO)
        cell.alignment = Alignment(horizontal='center')
        row += 2
        
        # LINHA 1: Receita, EBITDA, Lucro
        kpis_linha1 = [
            ("💰 RECEITA BRUTA", f"R$ {receita/1000:,.0f} mil", self.estilos.AZUL_MEDIO),
            ("📈 EBITDA", f"R$ {ebitda/1000:,.0f} mil", self.estilos.VERDE),
            ("💵 LUCRO LÍQUIDO", f"R$ {lucro/1000:,.0f} mil", self.estilos.VERDE_ESCURO),
        ]
        
        col = 2
        for titulo, valor, cor in kpis_linha1:
            # Card background
            for r in range(row, row + 3):
                for c in range(col, col + 2):
                    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=self.estilos.FUNDO_AZUL)
                    ws.cell(row=r, column=c).border = Border(
                        left=Side(style='medium', color=cor),
                        right=Side(style='medium', color=cor),
                        top=Side(style='medium', color=cor),
                        bottom=Side(style='medium', color=cor)
                    )
            
            # Título do KPI
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            cell = ws.cell(row=row, column=col)
            cell.value = titulo
            cell.font = Font(name='Calibri', size=9, bold=True, color=cor)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Valor do KPI
            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
            cell = ws.cell(row=row+1, column=col)
            cell.value = valor
            cell.font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            col += 2
        
        row += 4
        
        # LINHA 2: Margens
        kpis_linha2 = [
            ("📊 MARGEM EBITDA", f"{margem_ebitda*100:.1f}%", self.estilos.AZUL_MEDIO),
            ("📊 MARGEM LÍQUIDA", f"{margem_liq*100:.1f}%", self.estilos.VERDE),
            ("🎁 DIVIDENDOS", f"R$ {dividendos/1000:,.0f} mil", self.estilos.VERDE_ESCURO),
        ]
        
        col = 2
        for titulo, valor, cor in kpis_linha2:
            for r in range(row, row + 3):
                for c in range(col, col + 2):
                    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=self.estilos.FUNDO_VERDE)
                    ws.cell(row=r, column=c).border = Border(
                        left=Side(style='medium', color=cor),
                        right=Side(style='medium', color=cor),
                        top=Side(style='medium', color=cor),
                        bottom=Side(style='medium', color=cor)
                    )
            
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            cell = ws.cell(row=row, column=col)
            cell.value = titulo
            cell.font = Font(name='Calibri', size=9, bold=True, color=cor)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
            cell = ws.cell(row=row+1, column=col)
            cell.value = valor
            cell.font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            col += 2
        
        row += 4
        
        # LINHA 3: Operacionais
        kpis_linha3 = [
            ("🏥 SESSÕES/ANO", f"{sessoes:,.0f}", self.estilos.LARANJA),
            ("🎯 PONTO EQUILÍBRIO", f"R$ {pe_medio/1000:,.0f} mil/mês", self.estilos.LARANJA),
            ("📊 OCUPAÇÃO", f"{ocup_media*100:.0f}%", self.estilos.LARANJA),
        ]
        
        col = 2
        for titulo, valor, cor in kpis_linha3:
            for r in range(row, row + 3):
                for c in range(col, col + 2):
                    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=self.estilos.FUNDO_AMARELO)
                    ws.cell(row=r, column=c).border = Border(
                        left=Side(style='medium', color=cor),
                        right=Side(style='medium', color=cor),
                        top=Side(style='medium', color=cor),
                        bottom=Side(style='medium', color=cor)
                    )
            
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            cell = ws.cell(row=row, column=col)
            cell.value = titulo
            cell.font = Font(name='Calibri', size=9, bold=True, color=cor)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
            cell = ws.cell(row=row+1, column=col)
            cell.value = valor
            cell.font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            col += 2
        
        row += 5
        
        # =====================================================================
        # RESUMO RÁPIDO
        # =====================================================================
        ws.merge_cells(f'B{row}:G{row}')
        cell = ws.cell(row=row, column=2)
        cell.value = "RESUMO DO EXERCÍCIO"
        cell.font = Font(name='Calibri', size=12, bold=True, color=self.estilos.AZUL_ESCURO)
        cell.alignment = Alignment(horizontal='center')
        row += 2
        
        # Tabela resumo
        resumo = [
            ("Faturamento Bruto Anual", f"R$ {receita:,.0f}"),
            ("Impostos e Deduções", f"R$ {receita - receita_liq:,.0f}"),
            ("Custos e Despesas", f"R$ {receita_liq - lucro:,.0f}"),
            ("Resultado Líquido", f"R$ {lucro:,.0f}"),
            ("Margem de Segurança", f"{ms_media*100:.1f}%"),
        ]
        
        for label, valor in resumo:
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = Font(name='Calibri', size=10)
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            
            ws.cell(row=row, column=5).value = valor
            ws.cell(row=row, column=5).font = Font(name='Calibri', size=10, bold=True)
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5).border = self.estilos.borda_fina()
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
            
            row += 1
        
        row += 2
        
        # =====================================================================
        # RODAPÉ
        # =====================================================================
        for c in range(2, 8):
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=self.estilos.AZUL_ESCURO)
        ws.row_dimensions[row].height = 4
        
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        cell = ws.cell(row=row, column=2)
        cell.value = f"Documento gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        cell.font = Font(name='Calibri', size=9, italic=True, color=self.estilos.CINZA_MEDIO)
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        cell = ws.cell(row=row, column=2)
        cell.value = "Budget Engine - Sistema de Orçamento Empresarial"
        cell.font = Font(name='Calibri', size=9, color=self.estilos.CINZA_MEDIO)
        cell.alignment = Alignment(horizontal='center')
        
        return ws

    # =========================================================================
    # ABA 2: RESUMO EXECUTIVO
    # =========================================================================
    def criar_resumo_executivo(self):
        ws = self.wb.create_sheet("Resumo Executivo")
        self._set_larguras_padrao(ws)
        
        ws.merge_cells('B2:O2')
        ws['B2'] = "📊 RESUMO EXECUTIVO"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        
        # Evolução mensal
        row = self._criar_titulo_secao(ws, row, "EVOLUÇÃO MENSAL")
        row = self._criar_cabecalho_mensal(ws, row)
        
        row = self._escrever_linha(ws, row, "Receita Bruta", self.dre.get('Receita Bruta Total', [0]*12), nivel=0)
        row = self._escrever_linha(ws, row, "Receita Líquida", self.dre.get('Receita Líquida', [0]*12), nivel=1)
        row = self._escrever_linha(ws, row, "EBITDA", self.dre.get('EBITDA', [0]*12), is_resultado=True)
        
        margens = [self.dre.get('EBITDA', [0]*12)[m] / self.dre.get('Receita Líquida', [0]*12)[m] 
                   if self.dre.get('Receita Líquida', [0]*12)[m] > 0 else 0 for m in range(12)]
        row = self._escrever_linha(ws, row, "Margem EBITDA %", margens, nivel=1, is_percent=True)
        
        row = self._escrever_linha(ws, row, "Resultado Líquido", self.dre.get('Resultado Líquido', [0]*12), is_resultado=True)
        
        # Fluxo de caixa resumido
        row += 1
        row = self._criar_titulo_secao(ws, row, "FLUXO DE CAIXA")
        row = self._criar_cabecalho_mensal(ws, row)
        
        row = self._escrever_linha(ws, row, "Saldo Inicial", self.fc.get('Saldo Inicial', [0]*12), nivel=0)
        row = self._escrever_linha(ws, row, "Geração de Caixa", self.fc.get('(+/-) Variação', [0]*12), is_resultado=True)
        row = self._escrever_linha(ws, row, "Saldo Final", self.fc.get('Saldo Final', [0]*12), is_total=True)
        
        return ws

    # =========================================================================
    # ABA 2: DRE COMPLETO
    # =========================================================================
    def criar_dre(self):
        ws = self.wb.create_sheet("DRE")
        self._set_larguras_padrao(ws)
        
        ws.merge_cells('B2:P2')
        ws['B2'] = "📋 DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        receita_total = sum(self.dre.get('Receita Bruta Total', [0]*12))
        
        # RECEITAS
        row = self._criar_titulo_secao(ws, row, "RECEITA OPERACIONAL BRUTA", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        for servico in self.motor.servicos.keys():
            valores = self.dre.get(servico, [0]*12)
            row = self._escrever_linha(ws, row, servico, valores, nivel=1, receita_ref=receita_total)
        
        row = self._escrever_linha(ws, row, "RECEITA BRUTA TOTAL", 
                                   self.dre.get('Receita Bruta Total', [0]*12), 
                                   is_total=True, receita_ref=receita_total)
        row += 1
        
        # DEDUÇÕES
        row = self._criar_titulo_secao(ws, row, "(-) DEDUÇÕES DA RECEITA", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        row = self._escrever_linha(ws, row, "Simples Nacional", 
                                   self.dre.get('(-) Simples Nacional', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Taxa Cartão de Crédito", 
                                   self.dre.get('(-) Taxa Cartão', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "TOTAL DEDUÇÕES", 
                                   self.dre.get('Total Deduções', [0]*12), is_total=True, receita_ref=receita_total)
        
        row = self._escrever_linha(ws, row, "= RECEITA LÍQUIDA", 
                                   self.dre.get('Receita Líquida', [0]*12), is_resultado=True, receita_ref=receita_total)
        row += 1
        
        # CUSTOS VARIÁVEIS
        row = self._criar_titulo_secao(ws, row, "(-) CUSTOS VARIÁVEIS", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        row = self._escrever_linha(ws, row, "Materiais de Consumo", 
                                   self.dre.get('(-) Materiais', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "TOTAL CUSTOS VARIÁVEIS", 
                                   self.dre.get('Total Custos Variáveis', [0]*12), is_total=True, receita_ref=receita_total)
        
        row = self._escrever_linha(ws, row, "= MARGEM DE CONTRIBUIÇÃO", 
                                   self.dre.get('Margem de Contribuição', [0]*12), is_resultado=True, receita_ref=receita_total)
        row += 1
        
        # CUSTOS FIXOS - PESSOAL
        row = self._criar_titulo_secao(ws, row, "(-) CUSTOS COM PESSOAL", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        custos_pessoal = [
            ("Fisioterapeutas (PJ)", '(-) Folha Fisioterapeutas'),
            ("Proprietários", '(-) Folha Proprietários'),
            ("Pró-Labore", '(-) Pró-Labore'),
            ("CLT + Encargos", '(-) Folha CLT + Encargos'),
        ]
        for label, key in custos_pessoal:
            valores = self.dre.get(key, [0]*12)
            if sum(valores) != 0:
                row = self._escrever_linha(ws, row, label, valores, nivel=1, receita_ref=receita_total)
        
        row = self._escrever_linha(ws, row, "SUBTOTAL PESSOAL", 
                                   self.dre.get('Subtotal Pessoal', [0]*12), is_total=True, receita_ref=receita_total)
        row += 1
        
        # DESPESAS OPERACIONAIS
        row = self._criar_titulo_secao(ws, row, "(-) DESPESAS OPERACIONAIS", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        row = self._escrever_linha(ws, row, "TOTAL DESPESAS OPERACIONAIS", 
                                   self.dre.get('Total Despesas Fixas', [0]*12), is_total=True, receita_ref=receita_total)
        
        row = self._escrever_linha(ws, row, "TOTAL CUSTOS FIXOS", 
                                   self.dre.get('Total Custos Fixos', [0]*12), is_total=True, receita_ref=receita_total)
        row += 1
        
        # RESULTADO
        row = self._criar_titulo_secao(ws, row, "RESULTADO", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        row = self._escrever_linha(ws, row, "= EBITDA", 
                                   self.dre.get('EBITDA', [0]*12), is_resultado=True, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "= RESULTADO LÍQUIDO", 
                                   self.dre.get('Resultado Líquido', [0]*12), is_resultado=True, receita_ref=receita_total)
        
        # Destinação
        row += 1
        row = self._escrever_linha(ws, row, "(-) Reserva Legal (5%)", 
                                   self.dre.get('(-) Reserva Legal', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "(-) Reserva Investimentos (20%)", 
                                   self.dre.get('(-) Reserva Investimentos', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "(-) Dividendos (30%)", 
                                   self.dre.get('(-) Dividendos Distribuídos', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "= LUCRO RETIDO", 
                                   self.dre.get('Lucro no Período', [0]*12), is_resultado=True, receita_ref=receita_total)
        
        return ws

    # =========================================================================
    # ABA 3: DESPESAS DETALHADAS (MÊS A MÊS)
    # =========================================================================
    def criar_despesas_detalhadas(self):
        ws = self.wb.create_sheet("Despesas Detalhadas")
        self._set_larguras_padrao(ws)
        
        ws.merge_cells('B2:P2')
        ws['B2'] = "📋 DESPESAS FIXAS E VARIÁVEIS DETALHADAS"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        receita_total = sum(self.dre.get('Receita Bruta Total', [0]*12))
        
        # CUSTOS VARIÁVEIS
        row = self._criar_titulo_secao(ws, row, "CUSTOS VARIÁVEIS", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        receitas = self.dre.get('Receita Bruta Total', [0]*12)
        materiais = self.dre.get('(-) Materiais', [0]*12)
        taxa_cartao = self.dre.get('(-) Taxa Cartão', [0]*12)
        
        row = self._escrever_linha(ws, row, "Materiais (4% s/ receita)", materiais, nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Taxa Cartão (3,36%)", taxa_cartao, nivel=1, receita_ref=receita_total)
        
        total_cv = [materiais[m] + taxa_cartao[m] for m in range(12)]
        row = self._escrever_linha(ws, row, "TOTAL CUSTOS VARIÁVEIS", total_cv, is_total=True, receita_ref=receita_total)
        
        # % sobre receita
        pct_cv = [total_cv[m] / receitas[m] if receitas[m] > 0 else 0 for m in range(12)]
        row = self._escrever_linha(ws, row, "% sobre Receita", pct_cv, nivel=1, is_percent=True)
        row += 1
        
        # PESSOAL
        row = self._criar_titulo_secao(ws, row, "CUSTOS COM PESSOAL", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        row = self._escrever_linha(ws, row, "Fisioterapeutas (PJ)", 
                                   self.dre.get('(-) Folha Fisioterapeutas', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Proprietários", 
                                   self.dre.get('(-) Folha Proprietários', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Pró-Labore", 
                                   self.dre.get('(-) Pró-Labore', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "CLT + Encargos", 
                                   self.dre.get('(-) Folha CLT + Encargos', [0]*12), nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "SUBTOTAL PESSOAL", 
                                   self.dre.get('Subtotal Pessoal', [0]*12), is_total=True, receita_ref=receita_total)
        row += 1
        
        # DESPESAS OPERACIONAIS DETALHADAS
        row = self._criar_titulo_secao(ws, row, "DESPESAS OPERACIONAIS", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        despesas = [
            ("Aluguel", '(-) Aluguel'),
            ("Energia Elétrica", '(-) Energia'),
            ("TV/Telefone/Internet", '(-) TV/Telefone/Internet'),
            ("Limpeza", '(-) Limpeza'),
            ("Manutenção", '(-) Manutenção'),
            ("Seguros", '(-) Seguros'),
            ("Sistema/Software", '(-) Sistema'),
            ("Contabilidade", '(-) Contabilidade'),
            ("Marketing", '(-) Marketing'),
            ("Serviços de Terceiros", '(-) Serviços Terceiros'),
            ("Cursos/Treinamentos", '(-) Cursos'),
        ]
        
        for label, key in despesas:
            valores = self.dre.get(key, [0]*12)
            if sum(valores) != 0:
                row = self._escrever_linha(ws, row, label, valores, nivel=1, receita_ref=receita_total)
        
        row = self._escrever_linha(ws, row, "TOTAL DESPESAS OPERACIONAIS", 
                                   self.dre.get('Total Despesas Fixas', [0]*12), is_total=True, receita_ref=receita_total)
        row += 1
        
        # RESUMO POR CATEGORIA
        row = self._criar_titulo_secao(ws, row, "RESUMO POR CATEGORIA", col_fim=16)
        row = self._criar_cabecalho_mensal(ws, row, incluir_av=True)
        
        # Calcular por categoria
        pessoal = self.dre.get('Subtotal Pessoal', [0]*12)
        ocupacao = [self.dre.get('(-) Aluguel', [0]*12)[m] for m in range(12)]
        utilidades = [self.dre.get('(-) Energia', [0]*12)[m] + self.dre.get('(-) TV/Telefone/Internet', [0]*12)[m] for m in range(12)]
        admin = [self.dre.get('(-) Contabilidade', [0]*12)[m] + self.dre.get('(-) Sistema', [0]*12)[m] + self.dre.get('(-) Seguros', [0]*12)[m] for m in range(12)]
        marketing = self.dre.get('(-) Marketing', [0]*12)
        manutencao = [self.dre.get('(-) Limpeza', [0]*12)[m] + self.dre.get('(-) Manutenção', [0]*12)[m] for m in range(12)]
        
        row = self._escrever_linha(ws, row, "Pessoal", pessoal, nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Ocupação", ocupacao, nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Utilidades", utilidades, nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Administrativas", admin, nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Marketing", marketing, nivel=1, receita_ref=receita_total)
        row = self._escrever_linha(ws, row, "Manutenção/Limpeza", manutencao, nivel=1, receita_ref=receita_total)
        
        total_fixos = self.dre.get('Total Custos Fixos', [0]*12)
        row = self._escrever_linha(ws, row, "TOTAL CUSTOS FIXOS", total_fixos, is_total=True, receita_ref=receita_total)
        
        return ws

    # =========================================================================
    # ABA 4: PONTO DE EQUILÍBRIO COMPLETO
    # =========================================================================
    def criar_ponto_equilibrio(self):
        ws = self.wb.create_sheet("Ponto Equilíbrio")
        self._set_larguras_padrao(ws)
        
        ws.merge_cells('B2:O2')
        ws['B2'] = "⚖️ ANÁLISE DE PONTO DE EQUILÍBRIO"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        
        # DADOS BASE
        row = self._criar_titulo_secao(ws, row, "DADOS PARA CÁLCULO DO PONTO DE EQUILÍBRIO")
        row = self._criar_cabecalho_mensal(ws, row)
        
        rec_liq = [self.pe_anual.meses[m].receita_liquida for m in range(12)]
        cv = [self.pe_anual.meses[m].custos_variaveis for m in range(12)]
        mc = [self.pe_anual.meses[m].margem_contribuicao for m in range(12)]
        pct_mc = [self.pe_anual.meses[m].pct_margem_contribuicao for m in range(12)]
        cf = [self.pe_anual.meses[m].custos_fixos for m in range(12)]
        ebitda = [self.pe_anual.meses[m].ebitda for m in range(12)]
        
        row = self._escrever_linha(ws, row, "Receita Líquida", rec_liq, nivel=0)
        row = self._escrever_linha(ws, row, "(-) Custos Variáveis", [-v for v in cv], nivel=1)
        row = self._escrever_linha(ws, row, "= Margem Contribuição", mc, is_resultado=True)
        row = self._escrever_linha(ws, row, "% Margem Contribuição", pct_mc, nivel=1, is_percent=True)
        row = self._escrever_linha(ws, row, "(-) Custos Fixos", [-v for v in cf], nivel=1)
        row = self._escrever_linha(ws, row, "= EBITDA", ebitda, is_resultado=True)
        row += 1
        
        # PONTOS DE EQUILÍBRIO
        row = self._criar_titulo_secao(ws, row, "PONTOS DE EQUILÍBRIO")
        row = self._criar_cabecalho_mensal(ws, row)
        
        pe_contabil = [self.pe_anual.meses[m].pe_contabil for m in range(12)]
        pe_ociosidade = [self.pe_anual.meses[m].pe_com_ociosidade for m in range(12)]
        pe_sessoes = [self.pe_anual.meses[m].pe_sessoes for m in range(12)]
        pe_horas = [self.pe_anual.meses[m].pe_horas for m in range(12)]
        
        row = self._escrever_linha(ws, row, "PE Contábil (R$)", pe_contabil, nivel=0)
        row = self._escrever_linha(ws, row, "PE c/ Ociosidade (R$)", pe_ociosidade, nivel=1)
        row = self._escrever_linha(ws, row, "PE em Sessões", pe_sessoes, nivel=1)
        row = self._escrever_linha(ws, row, "PE em Horas", pe_horas, nivel=1)
        row += 1
        
        # REALIZADO vs PE
        row = self._criar_titulo_secao(ws, row, "REALIZADO vs PONTO DE EQUILÍBRIO")
        row = self._criar_cabecalho_mensal(ws, row)
        
        total_sessoes = [self.pe_anual.meses[m].total_sessoes for m in range(12)]
        total_horas = [self.pe_anual.meses[m].demanda_horas for m in range(12)]
        
        row = self._escrever_linha(ws, row, "Receita Líquida Realizada", rec_liq, nivel=0)
        row = self._escrever_linha(ws, row, "Ponto de Equilíbrio", pe_contabil, nivel=0)
        
        folga_rs = [rec_liq[m] - pe_contabil[m] for m in range(12)]
        row = self._escrever_linha(ws, row, "Folga (R$)", folga_rs, is_resultado=True)
        
        row += 1
        row = self._escrever_linha(ws, row, "Sessões Realizadas", total_sessoes, nivel=0)
        row = self._escrever_linha(ws, row, "Sessões PE", pe_sessoes, nivel=0)
        
        folga_sessoes = [total_sessoes[m] - pe_sessoes[m] for m in range(12)]
        row = self._escrever_linha(ws, row, "Folga (Sessões)", folga_sessoes, is_resultado=True)
        row += 1
        
        # INDICADORES DE RISCO
        row = self._criar_titulo_secao(ws, row, "INDICADORES DE RISCO OPERACIONAL")
        row = self._criar_cabecalho_mensal(ws, row)
        
        ms_valor = [self.pe_anual.meses[m].margem_seguranca_valor for m in range(12)]
        ms_pct = [self.pe_anual.meses[m].margem_seguranca_pct for m in range(12)]
        gao = [self.pe_anual.meses[m].gao for m in range(12)]
        
        row = self._escrever_linha(ws, row, "Margem Segurança (R$)", ms_valor, nivel=0)
        row = self._escrever_linha(ws, row, "Margem Segurança (%)", ms_pct, nivel=0, is_percent=True)
        row = self._escrever_linha(ws, row, "GAO (Alavancagem)", gao, nivel=0)
        
        # Interpretação
        row += 2
        ws.cell(row=row, column=2).value = "📖 INTERPRETAÇÃO:"
        ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True)
        row += 1
        
        notas = [
            "• Margem de Segurança: Quanto a receita pode cair antes de ter prejuízo. Ideal: >20%",
            "• GAO: Se a receita subir 1%, o lucro sobe GAO%. Alto GAO = maior risco e maior potencial",
            f"• GAO médio de {sum(gao)/12:.2f}x indica {'alta' if sum(gao)/12 > 3 else 'moderada'} sensibilidade do lucro à receita",
        ]
        for nota in notas:
            ws.cell(row=row, column=2).value = nota
            ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, color=self.estilos.CINZA_ESCURO)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=15)
            row += 1
        
        return ws

    # =========================================================================
    # ABA 5: TAXA DE OCUPAÇÃO COMPLETA
    # =========================================================================
    def criar_taxa_ocupacao(self):
        ws = self.wb.create_sheet("Taxa Ocupação")
        self._set_larguras_padrao(ws)
        
        ws.merge_cells('B2:O2')
        ws['B2'] = "📊 ANÁLISE DE CAPACIDADE E OCUPAÇÃO"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        
        # CAPACIDADE
        row = self._criar_titulo_secao(ws, row, "CAPACIDADE INSTALADA (HORAS)")
        row = self._criar_cabecalho_mensal(ws, row)
        
        cap_prof = [self.ocupacao.meses[m].capacidade_profissional for m in range(12)]
        cap_sala = [self.ocupacao.meses[m].capacidade_sala for m in range(12)]
        
        row = self._escrever_linha(ws, row, "Capacidade Profissional (h)", cap_prof, nivel=0)
        row = self._escrever_linha(ws, row, "Capacidade Sala (h)", cap_sala, nivel=0)
        
        # Capacidade limitante
        cap_min = [min(cap_prof[m], cap_sala[m]) for m in range(12)]
        row = self._escrever_linha(ws, row, "Capacidade Limitante (h)", cap_min, is_total=True)
        row += 1
        
        # DEMANDA
        row = self._criar_titulo_secao(ws, row, "DEMANDA (HORAS NECESSÁRIAS)")
        row = self._criar_cabecalho_mensal(ws, row)
        
        dem_prof = [self.ocupacao.meses[m].demanda_profissional for m in range(12)]
        dem_sala = [self.ocupacao.meses[m].demanda_sala for m in range(12)]
        
        row = self._escrever_linha(ws, row, "Demanda Profissional (h)", dem_prof, nivel=0)
        row = self._escrever_linha(ws, row, "Demanda Sala (h)", dem_sala, nivel=0)
        row += 1
        
        # OCIOSIDADE (calculada)
        row = self._criar_titulo_secao(ws, row, "OCIOSIDADE (HORAS NÃO UTILIZADAS)")
        row = self._criar_cabecalho_mensal(ws, row)
        
        ocio_prof = [cap_prof[m] - dem_prof[m] for m in range(12)]
        ocio_sala = [cap_sala[m] - dem_sala[m] for m in range(12)]
        
        row = self._escrever_linha(ws, row, "Ociosidade Profissional (h)", ocio_prof, nivel=1)
        row = self._escrever_linha(ws, row, "Ociosidade Sala (h)", ocio_sala, nivel=1)
        row += 1
        
        # TAXAS DE OCUPAÇÃO
        row = self._criar_titulo_secao(ws, row, "TAXAS DE OCUPAÇÃO")
        row = self._criar_cabecalho_mensal(ws, row)
        
        taxa_prof = [self.ocupacao.meses[m].taxa_ocupacao_profissional for m in range(12)]
        taxa_sala = [self.ocupacao.meses[m].taxa_ocupacao_sala for m in range(12)]
        
        row = self._escrever_linha(ws, row, "Taxa Ocupação Profissional", taxa_prof, nivel=0, is_percent=True)
        row = self._escrever_linha(ws, row, "Taxa Ocupação Sala", taxa_sala, nivel=0, is_percent=True)
        
        # Taxa efetiva (menor das duas)
        taxa_efetiva = [min(taxa_prof[m], taxa_sala[m]) for m in range(12)]
        row = self._escrever_linha(ws, row, "Taxa Ocupação Efetiva", taxa_efetiva, is_resultado=True, is_percent=True)
        row += 1
        
        # GARGALO
        row = self._criar_titulo_secao(ws, row, "ANÁLISE DE GARGALO")
        row = self._criar_cabecalho_mensal(ws, row)
        
        gargalos = [self.ocupacao.meses[m].gargalo for m in range(12)]
        
        # Linha de gargalo
        cell = ws.cell(row=row, column=2)
        cell.value = "Gargalo do Mês"
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.border = self.estilos.borda_fina()
        
        for i, gargalo in enumerate(gargalos):
            cell = ws.cell(row=row, column=3+i)
            cell.value = gargalo or "-"
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.estilos.borda_fina()
            if gargalo == "Sala":
                cell.fill = PatternFill('solid', fgColor=self.estilos.FUNDO_VERMELHO)
            elif gargalo == "Profissional":
                cell.fill = PatternFill('solid', fgColor=self.estilos.FUNDO_AMARELO)
        
        row += 2
        
        # Resumo
        gargalo_predominante = self.ocupacao.gargalo_predominante or "Não identificado"
        ws.merge_cells(f'B{row}:O{row}')
        cell = ws.cell(row=row, column=2)
        cell.value = f"⚠️ GARGALO PREDOMINANTE: {gargalo_predominante.upper()}"
        cell.font = Font(name='Calibri', size=14, bold=True, color=self.estilos.BRANCO)
        cell.fill = PatternFill('solid', fgColor=self.estilos.LARANJA)
        cell.alignment = Alignment(horizontal='center')
        
        row += 2
        
        # Recomendações
        ws.cell(row=row, column=2).value = "💡 RECOMENDAÇÕES:"
        ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True)
        row += 1
        
        media_sala = sum(taxa_sala) / 12
        media_prof = sum(taxa_prof) / 12
        
        recomendacoes = []
        if media_sala > 0.90:
            recomendacoes.append("• Ocupação de sala >90%: Considerar ampliação do espaço físico ou extensão de horário")
        if media_prof > 0.85:
            recomendacoes.append("• Ocupação profissional >85%: Considerar contratação de novos fisioterapeutas")
        if media_sala < 0.60:
            recomendacoes.append("• Ocupação de sala <60%: Há espaço ocioso, focar em captação de pacientes")
        if media_prof < 0.60:
            recomendacoes.append("• Ocupação profissional <60%: Revisar escalas ou redistribuir demanda")
        
        if not recomendacoes:
            recomendacoes.append("• Ocupação em níveis adequados (60-85%)")
        
        for rec in recomendacoes:
            ws.cell(row=row, column=2).value = rec
            ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, color=self.estilos.CINZA_ESCURO)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=15)
            row += 1
        
        return ws

    # =========================================================================
    # ABA 6: TDABC COMPLETO E EXPANDIDO
    # =========================================================================
    def criar_tdabc(self):
        ws = self.wb.create_sheet("TDABC")
        
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 18
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[col].width = 12
        
        ws.merge_cells('B2:M2')
        ws['B2'] = "🏆 CUSTEIO BASEADO EM ATIVIDADES (TDABC)"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        
        # Calcular dados anuais completos
        servicos_data = []
        for servico in self.motor.servicos.keys():
            sessoes = 0
            horas = 0
            m2 = 0
            receita = 0
            cv = 0
            overhead = 0
            
            for m in range(12):
                rateio = self.tdabc.meses[m].rateios.get(servico)
                lucro_obj = self.tdabc.meses[m].lucros.get(servico)
                
                if rateio:
                    sessoes += rateio.sessoes
                    horas += rateio.horas_sala
                    m2 += rateio.m2_alocado  # m2 é fixo, mas somamos para média
                    receita += rateio.receita
                
                if lucro_obj:
                    cv += lucro_obj.custos_variaveis_rateados
                    overhead += lucro_obj.overhead_rateado
            
            lucro = receita - cv - overhead
            m2_medio = m2 / 12 if m2 > 0 else 0
            
            servicos_data.append({
                'servico': servico,
                'sessoes': sessoes,
                'horas': horas,
                'm2': m2_medio,
                'receita': receita,
                'cv': cv,
                'overhead': overhead,
                'lucro': lucro,
                'margem': lucro / receita if receita > 0 else 0,
                'lucro_hora': lucro / horas if horas > 0 else 0,
                'lucro_m2': lucro / m2_medio if m2_medio > 0 else 0,
                'lucro_sessao': lucro / sessoes if sessoes > 0 else 0,
                'receita_hora': receita / horas if horas > 0 else 0,
                'receita_m2': receita / m2_medio if m2_medio > 0 else 0,
                'receita_sessao': receita / sessoes if sessoes > 0 else 0,
            })
        
        # Ordenar por lucro/hora (principal métrica)
        servicos_data.sort(key=lambda x: -x['lucro_hora'])
        rankings = ['🥇', '🥈', '🥉', '4º', '5º', '6º']
        
        # =====================================================================
        # TABELA 1: RESUMO ANUAL
        # =====================================================================
        row = self._criar_titulo_secao(ws, row, "RESUMO ANUAL POR SERVIÇO", col_fim=13)
        
        headers = ["Serviço", "Sessões", "Horas", "m²", "Receita", "CV", "Overhead", "Lucro ABC", "Margem%", "Rank"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        for i, srv in enumerate(servicos_data):
            ws.cell(row=row, column=2).value = srv['servico']
            ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            
            valores = [
                (srv['sessoes'], '#,##0'),
                (srv['horas'], '#,##0'),
                (srv['m2'], '0.0'),
                (srv['receita'], '#,##0'),
                (srv['cv'], '#,##0'),
                (srv['overhead'], '#,##0'),
                (srv['lucro'], '#,##0'),
            ]
            
            for j, (val, fmt) in enumerate(valores):
                cell = ws.cell(row=row, column=3+j)
                cell.value = val
                cell.number_format = fmt
                cell.border = self.estilos.borda_fina()
            
            # Margem com cor
            cell = ws.cell(row=row, column=10)
            cell.value = srv['margem']
            cell.number_format = '0.0%'
            cell.border = self.estilos.borda_fina()
            if srv['margem'] >= 0.20:
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.estilos.VERDE)
            elif srv['margem'] >= 0.10:
                cell.font = Font(name='Calibri', size=10, color=self.estilos.LARANJA)
            else:
                cell.font = Font(name='Calibri', size=10, color=self.estilos.VERMELHO)
            
            # Ranking
            ws.cell(row=row, column=11).value = rankings[i] if i < len(rankings) else f'{i+1}º'
            ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=11).border = self.estilos.borda_fina()
            ws.cell(row=row, column=11).font = Font(name='Calibri', size=12, bold=True)
            
            row += 1
        
        # Total
        total_receita = sum(s['receita'] for s in servicos_data)
        total_lucro = sum(s['lucro'] for s in servicos_data)
        total_horas = sum(s['horas'] for s in servicos_data)
        total_m2 = sum(s['m2'] for s in servicos_data)
        
        ws.cell(row=row, column=2).value = "TOTAL"
        ws.cell(row=row, column=3).value = sum(s['sessoes'] for s in servicos_data)
        ws.cell(row=row, column=4).value = total_horas
        ws.cell(row=row, column=5).value = total_m2
        ws.cell(row=row, column=6).value = total_receita
        ws.cell(row=row, column=7).value = sum(s['cv'] for s in servicos_data)
        ws.cell(row=row, column=8).value = sum(s['overhead'] for s in servicos_data)
        ws.cell(row=row, column=9).value = total_lucro
        ws.cell(row=row, column=10).value = total_lucro / total_receita if total_receita > 0 else 0
        ws.cell(row=row, column=10).number_format = '0.0%'
        
        for c in range(2, 12):
            self._aplicar_estilo_total(ws.cell(row=row, column=c))
        
        row += 2
        
        # =====================================================================
        # TABELA 2: INDICADORES DE EFICIÊNCIA
        # =====================================================================
        row = self._criar_titulo_secao(ws, row, "INDICADORES DE EFICIÊNCIA (LUCRO)", col_fim=13)
        
        headers = ["Serviço", "R$/Hora", "R$/m²", "R$/Sessão", "Rec/Hora", "Rec/m²", "Rec/Sessão", "Rank"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        for i, srv in enumerate(servicos_data):
            ws.cell(row=row, column=2).value = srv['servico']
            ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            
            valores = [
                srv['lucro_hora'] if srv['horas'] > 0 else None,
                srv['lucro_m2'] if srv['m2'] > 0 else None,
                srv['lucro_sessao'],
                srv['receita_hora'] if srv['horas'] > 0 else None,
                srv['receita_m2'] if srv['m2'] > 0 else None,
                srv['receita_sessao'],
            ]
            
            for j, val in enumerate(valores):
                cell = ws.cell(row=row, column=3+j)
                if val is None:
                    cell.value = "N/A"
                    cell.font = Font(name='Calibri', size=9, italic=True, color=self.estilos.CINZA_MEDIO)
                else:
                    cell.value = val
                    cell.number_format = '#,##0.00'
                    # Colorir lucro/hora (principal métrica)
                    if j == 0:
                        if val >= 40:
                            cell.font = Font(name='Calibri', size=10, bold=True, color=self.estilos.VERDE)
                        elif val >= 20:
                            cell.font = Font(name='Calibri', size=10, color=self.estilos.LARANJA)
                        else:
                            cell.font = Font(name='Calibri', size=10, color=self.estilos.VERMELHO)
                cell.border = self.estilos.borda_fina()
            
            # Ranking (já ordenado por lucro/hora)
            ws.cell(row=row, column=9).value = rankings[i] if i < len(rankings) else f'{i+1}º'
            ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=9).border = self.estilos.borda_fina()
            ws.cell(row=row, column=9).font = Font(name='Calibri', size=12, bold=True)
            
            row += 1
        
        # Total/Média
        ws.cell(row=row, column=2).value = "MÉDIA/TOTAL"
        ws.cell(row=row, column=3).value = total_lucro / total_horas if total_horas > 0 else 0
        ws.cell(row=row, column=4).value = total_lucro / total_m2 if total_m2 > 0 else 0
        ws.cell(row=row, column=5).value = total_lucro / sum(s['sessoes'] for s in servicos_data)
        ws.cell(row=row, column=6).value = total_receita / total_horas if total_horas > 0 else 0
        ws.cell(row=row, column=7).value = total_receita / total_m2 if total_m2 > 0 else 0
        ws.cell(row=row, column=8).value = total_receita / sum(s['sessoes'] for s in servicos_data)
        
        for c in range(2, 10):
            self._aplicar_estilo_total(ws.cell(row=row, column=c))
            if c >= 3:
                ws.cell(row=row, column=c).number_format = '#,##0.00'
        
        row += 2
        
        # =====================================================================
        # TABELA 3: RATEIOS DE CUSTOS (detalhamento)
        # =====================================================================
        row = self._criar_titulo_secao(ws, row, "DETALHAMENTO DOS RATEIOS (MÊS 6 - EXEMPLO)", col_fim=13)
        
        headers = ["Serviço", "Rateio m²", "Rateio Sessões", "Rateio Receita", "Total Rateio", "% do Total"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        mes_exemplo = 5
        total_rateio = 0
        
        for servico in self.motor.servicos.keys():
            rateio = self.tdabc.meses[mes_exemplo].rateios.get(servico)
            if rateio:
                rateio_total = rateio.rateio_m2 + rateio.rateio_sessoes + rateio.rateio_receita
                total_rateio += rateio_total
                
                ws.cell(row=row, column=2).value = servico
                ws.cell(row=row, column=2).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=3).value = rateio.rateio_m2
                ws.cell(row=row, column=3).number_format = '#,##0'
                ws.cell(row=row, column=3).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=4).value = rateio.rateio_sessoes
                ws.cell(row=row, column=4).number_format = '#,##0'
                ws.cell(row=row, column=4).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=5).value = rateio.rateio_receita
                ws.cell(row=row, column=5).number_format = '#,##0'
                ws.cell(row=row, column=5).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=6).value = rateio_total
                ws.cell(row=row, column=6).number_format = '#,##0'
                ws.cell(row=row, column=6).border = self.estilos.borda_fina()
                ws.cell(row=row, column=6).font = Font(name='Calibri', size=10, bold=True)
                
                row += 1
        
        # Preencher % depois de ter o total
        row_start = row - len(self.motor.servicos)
        for i, servico in enumerate(self.motor.servicos.keys()):
            rateio = self.tdabc.meses[mes_exemplo].rateios.get(servico)
            if rateio:
                rateio_total = rateio.rateio_m2 + rateio.rateio_sessoes + rateio.rateio_receita
                cell = ws.cell(row=row_start + i, column=7)
                cell.value = rateio_total / total_rateio if total_rateio > 0 else 0
                cell.number_format = '0.0%'
                cell.border = self.estilos.borda_fina()
        
        row += 1
        
        # =====================================================================
        # TABELA 4: EVOLUÇÃO MENSAL DO LUCRO ABC
        # =====================================================================
        self._set_larguras_padrao(ws, col_descricao=18)
        
        row = self._criar_titulo_secao(ws, row, "EVOLUÇÃO MENSAL DO LUCRO ABC", col_fim=15)
        row = self._criar_cabecalho_mensal(ws, row)
        
        for servico in self.motor.servicos.keys():
            lucros_mes = []
            for m in range(12):
                lucro_obj = self.tdabc.meses[m].lucros.get(servico)
                rateio = self.tdabc.meses[m].rateios.get(servico)
                if lucro_obj and rateio:
                    lucro = rateio.receita - lucro_obj.custos_variaveis_rateados - lucro_obj.overhead_rateado
                else:
                    lucro = 0
                lucros_mes.append(lucro)
            
            row = self._escrever_linha(ws, row, servico, lucros_mes, nivel=1)
        
        # Total
        total_mes = []
        for m in range(12):
            total = 0
            for servico in self.motor.servicos.keys():
                lucro_obj = self.tdabc.meses[m].lucros.get(servico)
                rateio = self.tdabc.meses[m].rateios.get(servico)
                if lucro_obj and rateio:
                    total += rateio.receita - lucro_obj.custos_variaveis_rateados - lucro_obj.overhead_rateado
            total_mes.append(total)
        
        row = self._escrever_linha(ws, row, "TOTAL LUCRO ABC", total_mes, is_total=True)
        row += 1
        
        # =====================================================================
        # TABELA 5: LUCRO/HORA MENSAL
        # =====================================================================
        row = self._criar_titulo_secao(ws, row, "LUCRO POR HORA (R$/h) - EVOLUÇÃO MENSAL", col_fim=15)
        row = self._criar_cabecalho_mensal(ws, row)
        
        for servico in self.motor.servicos.keys():
            lucro_hora_mes = []
            for m in range(12):
                lucro_obj = self.tdabc.meses[m].lucros.get(servico)
                rateio = self.tdabc.meses[m].rateios.get(servico)
                if lucro_obj and rateio and rateio.horas_sala > 0:
                    lucro = rateio.receita - lucro_obj.custos_variaveis_rateados - lucro_obj.overhead_rateado
                    lucro_hora = lucro / rateio.horas_sala
                else:
                    lucro_hora = 0
                lucro_hora_mes.append(lucro_hora)
            
            row = self._escrever_linha(ws, row, servico, lucro_hora_mes, nivel=1)
        
        row += 1
        
        # =====================================================================
        # TABELA 6: MARGEM ABC MENSAL
        # =====================================================================
        row = self._criar_titulo_secao(ws, row, "MARGEM ABC (%) - EVOLUÇÃO MENSAL", col_fim=15)
        row = self._criar_cabecalho_mensal(ws, row)
        
        for servico in self.motor.servicos.keys():
            margens_mes = []
            for m in range(12):
                lucro_obj = self.tdabc.meses[m].lucros.get(servico)
                rateio = self.tdabc.meses[m].rateios.get(servico)
                if lucro_obj and rateio and rateio.receita > 0:
                    lucro = rateio.receita - lucro_obj.custos_variaveis_rateados - lucro_obj.overhead_rateado
                    margem = lucro / rateio.receita
                else:
                    margem = 0
                margens_mes.append(margem)
            
            row = self._escrever_linha(ws, row, servico, margens_mes, nivel=1, is_percent=True)
        
        row += 2
        
        # =====================================================================
        # INTERPRETAÇÃO
        # =====================================================================
        ws.cell(row=row, column=2).value = "📖 INTERPRETAÇÃO E RECOMENDAÇÕES:"
        ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=15)
        row += 1
        
        # Identificar melhor e pior (apenas serviços com horas > 0)
        servicos_com_sala = [s for s in servicos_data if s['horas'] > 0]
        servicos_sem_sala = [s for s in servicos_data if s['horas'] == 0]
        
        if servicos_com_sala:
            melhor = servicos_com_sala[0]
            pior = servicos_com_sala[-1]
        
        interpretacoes = []
        
        if servicos_com_sala:
            interpretacoes.append(f"• 🥇 MELHOR (R$/hora): {melhor['servico']} - R$ {melhor['lucro_hora']:.2f}/hora | Margem {melhor['margem']*100:.1f}%")
            interpretacoes.append(f"• ⚠️ REVISAR: {pior['servico']} - R$ {pior['lucro_hora']:.2f}/hora | Margem {pior['margem']*100:.1f}%")
        
        if servicos_sem_sala:
            melhor_sem_sala = max(servicos_sem_sala, key=lambda x: x['lucro_sessao'])
            interpretacoes.append(f"• 🏠 DOMICILIAR: {melhor_sem_sala['servico']} - R$ {melhor_sem_sala['lucro_sessao']:.2f}/sessão | Margem {melhor_sem_sala['margem']*100:.1f}%")
        
        interpretacoes.append(f"• Lucro médio: R$ {total_lucro/total_horas:.2f}/hora | R$ {total_lucro/sum(s['sessoes'] for s in servicos_data):.2f}/sessão")
        interpretacoes.append(f"• Recomendação: Priorizar serviços com maior R$/hora na alocação de agenda")
        
        for interp in interpretacoes:
            ws.cell(row=row, column=2).value = interp
            ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, color=self.estilos.CINZA_ESCURO)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=15)
            row += 1
        
        return ws

    # =========================================================================
    # ABA 7: FLUXO DE CAIXA E IMPOSTOS
    # =========================================================================
    def criar_fluxo_caixa(self):
        ws = self.wb.create_sheet("Fluxo de Caixa")
        self._set_larguras_padrao(ws)
        
        ws.merge_cells('B2:O2')
        ws['B2'] = "💰 FLUXO DE CAIXA E TRIBUTOS"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        
        # ENTRADAS
        row = self._criar_titulo_secao(ws, row, "(+) ENTRADAS DE CAIXA")
        row = self._criar_cabecalho_mensal(ws, row)
        
        for servico in self.motor.servicos.keys():
            key = f'(+) {servico}'
            valores = self.fc.get(key, [0]*12)
            if sum(valores) > 0:
                row = self._escrever_linha(ws, row, f"Receb. {servico}", valores, nivel=1)
        
        row = self._escrever_linha(ws, row, "TOTAL ENTRADAS", 
                                   self.fc.get('Total Entradas', [0]*12), is_total=True)
        row += 1
        
        # SAÍDAS
        row = self._criar_titulo_secao(ws, row, "(-) SAÍDAS DE CAIXA")
        row = self._criar_cabecalho_mensal(ws, row)
        
        saidas = [
            ("Folha Proprietários", '(-) Folha Proprietários'),
            ("Folha Fisioterapeutas", '(-) Folha Fisioterapeutas'),
            ("Folha CLT Líquida", '(-) Folha CLT Líquida'),
            ("INSS + FGTS", '(-) INSS + FGTS'),
            ("Pró-labore + INSS", '(-) Pró-labore + INSS'),
            ("DAS Simples Nacional", '(-) DAS Simples Nacional'),
            ("Despesas Operacionais", '(-) Despesas Operacionais'),
            ("Custos Financeiros Cartão", '(-) Custos Financeiros Cartão'),
            ("Dividendos", '(-) Distribuição Dividendos'),
        ]
        
        for label, key in saidas:
            valores = self.fc.get(key, [0]*12)
            if sum(valores) != 0:
                row = self._escrever_linha(ws, row, label, valores, nivel=1)
        
        row = self._escrever_linha(ws, row, "TOTAL SAÍDAS", 
                                   self.fc.get('Total Saídas', [0]*12), is_total=True)
        row += 1
        
        # MOVIMENTAÇÃO
        row = self._criar_titulo_secao(ws, row, "MOVIMENTAÇÃO DE CAIXA")
        row = self._criar_cabecalho_mensal(ws, row)
        
        row = self._escrever_linha(ws, row, "Saldo Inicial", self.fc.get('Saldo Inicial', [0]*12), nivel=0)
        row = self._escrever_linha(ws, row, "(+/-) Geração de Caixa", self.fc.get('(+/-) Variação', [0]*12), is_resultado=True)
        row = self._escrever_linha(ws, row, "= SALDO FINAL", self.fc.get('Saldo Final', [0]*12), is_total=True)
        row += 1
        
        # CRONOGRAMA DE IMPOSTOS
        row = self._criar_titulo_secao(ws, row, "CRONOGRAMA DE IMPOSTOS")
        row = self._criar_cabecalho_mensal(ws, row)
        
        das = self.fc.get('(-) DAS Simples Nacional', [0]*12)
        inss_fgts = self.fc.get('(-) INSS + FGTS', [0]*12)
        inss_socio = self.fc.get('(-) Pró-labore + INSS', [0]*12)
        
        row = self._escrever_linha(ws, row, "DAS Simples Nacional", das, nivel=1)
        row = self._escrever_linha(ws, row, "INSS + FGTS (CLT)", inss_fgts, nivel=1)
        row = self._escrever_linha(ws, row, "INSS Sócio (Pró-labore)", inss_socio, nivel=1)
        
        total_impostos = [das[m] + inss_fgts[m] + inss_socio[m] for m in range(12)]
        row = self._escrever_linha(ws, row, "TOTAL IMPOSTOS/ENCARGOS", total_impostos, is_total=True)
        
        # % sobre receita
        receitas = self.dre.get('Receita Bruta Total', [0]*12)
        pct_impostos = [abs(total_impostos[m]) / receitas[m] if receitas[m] > 0 else 0 for m in range(12)]
        row = self._escrever_linha(ws, row, "% sobre Receita", pct_impostos, nivel=1, is_percent=True)
        
        return ws

    # =========================================================================
    # ABA 8: EQUIPE E FOLHA
    # =========================================================================
    def criar_equipe(self):
        ws = self.wb.create_sheet("Equipe")
        
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 18
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 12
        
        ws.merge_cells('B2:I2')
        ws['B2'] = "👥 EQUIPE E FOLHA DE PAGAMENTO"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        
        # FISIOTERAPEUTAS
        row = self._criar_titulo_secao(ws, row, "FISIOTERAPEUTAS", col_fim=9)
        
        headers = ["Nome", "Cargo", "Nível", "H/Sem", "Sessões", "Serviços"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        total_horas = 0
        total_sessoes = 0
        
        for nome, fisio in self.motor.fisioterapeutas.items():
            if fisio.ativo:
                horas = sum(fisio.escala_semanal.values())
                sessoes = sum(fisio.sessoes_por_servico.values())
                servicos = ', '.join([s[:3] for s, q in fisio.sessoes_por_servico.items() if q > 0])
                
                total_horas += horas
                total_sessoes += sessoes
                
                ws.cell(row=row, column=2).value = nome
                ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
                ws.cell(row=row, column=2).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=3).value = fisio.cargo
                ws.cell(row=row, column=3).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=4).value = fisio.nivel
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=4).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=5).value = horas
                ws.cell(row=row, column=5).number_format = '0.0'
                ws.cell(row=row, column=5).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=6).value = sessoes
                ws.cell(row=row, column=6).number_format = '#,##0'
                ws.cell(row=row, column=6).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=7).value = servicos
                ws.cell(row=row, column=7).border = self.estilos.borda_fina()
                
                row += 1
        
        # Total fisios
        ws.cell(row=row, column=2).value = "TOTAL FISIOTERAPEUTAS"
        ws.cell(row=row, column=5).value = total_horas
        ws.cell(row=row, column=6).value = total_sessoes
        for c in range(2, 8):
            self._aplicar_estilo_total(ws.cell(row=row, column=c))
        row += 2
        
        # FUNCIONÁRIOS CLT
        row = self._criar_titulo_secao(ws, row, "FUNCIONÁRIOS CLT", col_fim=9)
        
        headers = ["Nome", "Cargo", "Vínculo", "Salário"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        total_salario = 0
        for nome, func in self.motor.funcionarios_clt.items():
            if func.ativo:
                total_salario += func.salario_base
                
                ws.cell(row=row, column=2).value = nome
                ws.cell(row=row, column=2).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=3).value = func.cargo
                ws.cell(row=row, column=3).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=4).value = func.tipo_vinculo.upper()
                ws.cell(row=row, column=4).border = self.estilos.borda_fina()
                
                ws.cell(row=row, column=5).value = func.salario_base
                ws.cell(row=row, column=5).number_format = '#,##0'
                ws.cell(row=row, column=5).border = self.estilos.borda_fina()
                
                row += 1
        
        # Total CLT
        ws.cell(row=row, column=2).value = "TOTAL CLT"
        ws.cell(row=row, column=5).value = total_salario
        ws.cell(row=row, column=5).number_format = '#,##0'
        for c in range(2, 6):
            self._aplicar_estilo_total(ws.cell(row=row, column=c))
        row += 2
        
        # SÓCIOS
        row = self._criar_titulo_secao(ws, row, "SÓCIOS", col_fim=9)
        
        headers = ["Nome", "Participação", "Capital", "Pró-labore", "Dividendos/Ano"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        div_por_socio = self.dividendos.get('dividendos_por_socio', {})
        
        for nome, socio in self.motor.socios_prolabore.items():
            ws.cell(row=row, column=2).value = nome
            ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=3).value = socio.participacao
            ws.cell(row=row, column=3).number_format = '0.0%'
            ws.cell(row=row, column=3).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=4).value = socio.capital
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=4).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=5).value = socio.prolabore
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=5).border = self.estilos.borda_fina()
            
            div = div_por_socio.get(nome, {}).get('total_anual', 0)
            ws.cell(row=row, column=6).value = div
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=6).border = self.estilos.borda_fina()
            ws.cell(row=row, column=6).font = Font(name='Calibri', size=10, bold=True, color=self.estilos.VERDE)
            
            row += 1
        
        return ws

    # =========================================================================
    # ABA 9: SIMPLES NACIONAL E DIVIDENDOS
    # =========================================================================
    def criar_simples_dividendos(self):
        ws = self.wb.create_sheet("Simples e Dividendos")
        self._set_larguras_padrao(ws, col_descricao=20)
        
        ws.merge_cells('B2:O2')
        ws['B2'] = "🧾 SIMPLES NACIONAL E DIVIDENDOS"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        
        # SIMPLES NACIONAL
        row = self._criar_titulo_secao(ws, row, "PROJEÇÃO SIMPLES NACIONAL")
        
        headers = ["Mês", "Receita", "RBT12", "Fator R", "Anexo", "Alíq. Efet.", "DAS"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        total_das = 0
        for proj in self.simples.get('projecao_pj', []):
            mes = proj.get('mes', 1)
            
            ws.cell(row=row, column=2).value = self.MESES[mes-1]
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=3).value = proj.get('receita_mensal', 0)
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=3).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=4).value = proj.get('rbt12', 0)
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=4).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=5).value = proj.get('fator_r', 0)
            ws.cell(row=row, column=5).number_format = '0.00%'
            ws.cell(row=row, column=5).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=6).value = proj.get('anexo', '')
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=7).value = proj.get('aliquota_efetiva', 0)
            ws.cell(row=row, column=7).number_format = '0.00%'
            ws.cell(row=row, column=7).border = self.estilos.borda_fina()
            
            das = proj.get('das', 0)
            total_das += das
            ws.cell(row=row, column=8).value = das
            ws.cell(row=row, column=8).number_format = '#,##0'
            ws.cell(row=row, column=8).border = self.estilos.borda_fina()
            ws.cell(row=row, column=8).font = Font(name='Calibri', size=10, color=self.estilos.VERMELHO)
            
            row += 1
        
        # Total
        ws.cell(row=row, column=2).value = "TOTAL"
        ws.cell(row=row, column=8).value = total_das
        ws.cell(row=row, column=8).number_format = '#,##0'
        for c in range(2, 9):
            self._aplicar_estilo_total(ws.cell(row=row, column=c))
        
        row += 2
        
        # DIVIDENDOS
        row = self._criar_titulo_secao(ws, row, "CRONOGRAMA DE DIVIDENDOS")
        
        headers = ["Período", "Meses", "Lucro Acum.", "Dividendo", "Mês Pgto"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        for periodo in self.dividendos.get('dividendos_periodo', []):
            ws.cell(row=row, column=2).value = periodo.get('periodo', '')
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=3).value = f"{self.MESES[periodo.get('inicio', 1)-1]} a {self.MESES[periodo.get('fim', 3)-1]}"
            ws.cell(row=row, column=3).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=4).value = periodo.get('lucro_acumulado', 0)
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=4).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=5).value = periodo.get('dividendo', 0)
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=5).border = self.estilos.borda_fina()
            ws.cell(row=row, column=5).font = Font(name='Calibri', size=10, bold=True, color=self.estilos.VERDE)
            
            mes_pag = periodo.get('mes_pagamento', 1)
            ws.cell(row=row, column=6).value = self.MESES[mes_pag-1] if mes_pag <= 12 else ''
            ws.cell(row=row, column=6).border = self.estilos.borda_fina()
            
            row += 1
        
        # Total dividendos
        total_div = sum(self.dividendos.get('cronograma', [0]*12))
        ws.cell(row=row, column=2).value = "TOTAL ANUAL"
        ws.cell(row=row, column=5).value = total_div
        ws.cell(row=row, column=5).number_format = '#,##0'
        for c in range(2, 7):
            self._aplicar_estilo_total(ws.cell(row=row, column=c))
        
        return ws

    # =========================================================================
    # ABA 10: PREMISSAS
    # =========================================================================
    def criar_premissas(self):
        ws = self.wb.create_sheet("Premissas")
        
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        ws.merge_cells('B2:E2')
        ws['B2'] = "⚙️ PREMISSAS DO ORÇAMENTO"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.estilos.AZUL_ESCURO)
        
        row = 4
        
        # Operacionais
        row = self._criar_titulo_secao(ws, row, "PREMISSAS OPERACIONAIS", col_fim=5)
        row += 1
        
        premissas = [
            ("Horas funcionamento/dia", f"{self.motor.operacional.horas_atendimento_dia}h"),
            ("Dias úteis/mês", str(self.motor.operacional.dias_uteis_mes)),
            ("Número de salas", str(self.motor.operacional.num_salas)),
            ("Modelo tributário", self.motor.operacional.modelo_tributario),
            ("Taxa cartão", "3,36%"),
            ("Taxa materiais", "4,00%"),
        ]
        
        for param, valor in premissas:
            ws.cell(row=row, column=2).value = param
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            ws.cell(row=row, column=3).value = valor
            ws.cell(row=row, column=3).font = Font(name='Calibri', size=10, bold=True)
            ws.cell(row=row, column=3).border = self.estilos.borda_fina()
            row += 1
        
        row += 1
        
        # Serviços
        row = self._criar_titulo_secao(ws, row, "TABELA DE SERVIÇOS", col_fim=5)
        
        headers = ["Serviço", "Valor 2026", "Duração", "Usa Sala"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=2+i)
            cell.value = h
            self._aplicar_estilo_cabecalho(cell)
        row += 1
        
        for nome, srv in self.motor.servicos.items():
            ws.cell(row=row, column=2).value = nome
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=3).value = srv.valor_2026
            ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=3).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=4).value = f"{srv.duracao_minutos} min"
            ws.cell(row=row, column=4).border = self.estilos.borda_fina()
            
            ws.cell(row=row, column=5).value = "Sim" if srv.usa_sala else "Não"
            ws.cell(row=row, column=5).border = self.estilos.borda_fina()
            
            row += 1
        
        row += 1
        
        # Despesas Fixas
        row = self._criar_titulo_secao(ws, row, "DESPESAS FIXAS MENSAIS", col_fim=5)
        row += 1
        
        for nome, desp in self.motor.despesas_fixas.items():
            ws.cell(row=row, column=2).value = nome
            ws.cell(row=row, column=2).border = self.estilos.borda_fina()
            ws.cell(row=row, column=3).value = desp.valor_mensal
            ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=3).border = self.estilos.borda_fina()
            row += 1
        
        return ws

    # =========================================================================
    # GERAÇÃO FINAL
    # =========================================================================
    def generate(self, filepath: str):
        """Gera o arquivo Excel completo"""
        self.criar_capa()
        self.criar_resumo_executivo()
        self.criar_dre()
        self.criar_despesas_detalhadas()
        self.criar_ponto_equilibrio()
        self.criar_taxa_ocupacao()
        self.criar_tdabc()
        self.criar_fluxo_caixa()
        self.criar_equipe()
        self.criar_simples_dividendos()
        self.criar_premissas()
        
        self.wb.save(filepath)
        return filepath


def exportar_budget_cliente(motor, filepath: str, cliente: str = None, filial: str = None):
    """Função de conveniência - usa dados do motor se não especificados"""
    exporter = ExcelBudgetExporter(motor, cliente, filial)
    return exporter.generate(filepath)
